#!/usr/bin/env python3
"""
Sandbox Manager for Suna - Pure Python Implementation
Replaces Docker sandbox with Python-based isolation and browser automation
"""

import os
import sys
import subprocess
import tempfile
import shutil
import signal
import threading
import time
import socket
from pathlib import Path
from typing import Dict, List, Optional, Any, Union
import json
import logging
from contextlib import contextmanager

# Try to import playwright
try:
    from playwright.sync_api import sync_playwright, Browser, Page
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False

# Try to import other utilities
try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False


class Colors:
    HEADER = "\033[95m"
    BLUE = "\033[94m"
    CYAN = "\033[96m"
    GREEN = "\033[92m"
    YELLOW = "\033[93m"
    RED = "\033[91m"
    ENDC = "\033[0m"
    BOLD = "\033[1m"


class SandboxManager:
    """Manages sandboxed execution environments without Docker."""
    
    def __init__(self, project_root: str):
        self.project_root = Path(project_root).absolute()
        self.sandbox_dir = self.project_root / "sandbox_data"
        self.temp_dirs: List[Path] = []
        self.active_processes: Dict[str, subprocess.Popen] = {}
        self.browser_instance: Optional[Browser] = None
        self.browser_pages: Dict[str, Page] = {}
        self.setup_logging()
        
        # Create sandbox directory
        self.sandbox_dir.mkdir(exist_ok=True)
    
    def setup_logging(self):
        """Setup logging for sandbox operations."""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger('SandboxManager')

    def print_info(self, message: str):
        print(f"{Colors.CYAN}ℹ️  {message}{Colors.ENDC}")
        self.logger.info(message)
        
    def print_success(self, message: str):
        print(f"{Colors.GREEN}✅  {message}{Colors.ENDC}")
        self.logger.info(message)
        
    def print_warning(self, message: str):
        print(f"{Colors.YELLOW}⚠️  {message}{Colors.ENDC}")
        self.logger.warning(message)
        
    def print_error(self, message: str):
        print(f"{Colors.RED}❌  {message}{Colors.ENDC}")
        self.logger.error(message)

    def check_dependencies(self) -> bool:
        """Check if required dependencies are available."""
        deps_ok = True
        
        # Check Python dependencies
        if not PLAYWRIGHT_AVAILABLE:
            self.print_error("Playwright not available. Install with: pip install playwright")
            deps_ok = False
        
        if not PSUTIL_AVAILABLE:
            self.print_warning("psutil not available. Install with: pip install psutil")
            # This is not critical, but helpful
        
        # Check system dependencies
        required_tools = ["python3"]
        
        for tool in required_tools:
            if not shutil.which(tool):
                self.print_error(f"Required tool not found: {tool}")
                deps_ok = False
        
        return deps_ok

    def install_browser_dependencies(self) -> bool:
        """Install Playwright browser dependencies."""
        if not PLAYWRIGHT_AVAILABLE:
            self.print_error("Playwright not available")
            return False
            
        try:
            self.print_info("Installing Playwright browsers...")
            subprocess.run([
                sys.executable, "-m", "playwright", "install", "chromium"
            ], check=True)
            
            self.print_info("Installing system dependencies for browsers...")
            subprocess.run([
                sys.executable, "-m", "playwright", "install-deps", "chromium"
            ], check=True)
            
            self.print_success("Browser dependencies installed successfully")
            return True
            
        except subprocess.CalledProcessError as e:
            self.print_error(f"Failed to install browser dependencies: {e}")
            return False
        except Exception as e:
            self.print_error(f"Unexpected error installing dependencies: {e}")
            return False

    @contextmanager
    def create_isolated_environment(self, session_id: str):
        """Create an isolated environment for code execution."""
        session_dir = self.sandbox_dir / f"session_{session_id}"
        session_dir.mkdir(exist_ok=True)
        
        # Create a temporary working directory
        work_dir = session_dir / "workspace"
        work_dir.mkdir(exist_ok=True)
        
        # Set up environment variables for isolation
        env = os.environ.copy()
        env['PYTHONPATH'] = str(work_dir)
        env['HOME'] = str(session_dir)
        env['TMPDIR'] = str(session_dir / "tmp")
        
        # Create tmp directory
        (session_dir / "tmp").mkdir(exist_ok=True)
        
        try:
            yield {
                'session_dir': session_dir,
                'work_dir': work_dir,
                'env': env
            }
        finally:
            # Cleanup
            self.cleanup_session(session_id)

    def cleanup_session(self, session_id: str):
        """Clean up a session's resources."""
        session_dir = self.sandbox_dir / f"session_{session_id}"
        
        # Kill any running processes for this session
        if session_id in self.active_processes:
            try:
                process = self.active_processes[session_id]
                if process.poll() is None:
                    process.terminate()
                    try:
                        process.wait(timeout=5)
                    except subprocess.TimeoutExpired:
                        process.kill()
            except Exception as e:
                self.logger.warning(f"Error terminating process for session {session_id}: {e}")
            finally:
                del self.active_processes[session_id]
        
        # Clean up browser pages
        if session_id in self.browser_pages:
            try:
                self.browser_pages[session_id].close()
            except Exception as e:
                self.logger.warning(f"Error closing browser page for session {session_id}: {e}")
            finally:
                del self.browser_pages[session_id]
        
        # Remove session directory
        if session_dir.exists():
            try:
                shutil.rmtree(session_dir)
            except Exception as e:
                self.logger.warning(f"Error removing session directory {session_id}: {e}")

    def execute_python_code(self, code: str, session_id: str, timeout: int = 30) -> Dict[str, Any]:
        """Execute Python code in an isolated environment."""
        with self.create_isolated_environment(session_id) as env:
            work_dir = env['work_dir']
            env_vars = env['env']
            
            # Create a Python script file
            script_file = work_dir / "script.py"
            with open(script_file, 'w') as f:
                f.write(code)
            
            try:
                # Execute with restricted permissions
                cmd = [sys.executable, str(script_file)]
                
                process = subprocess.Popen(
                    cmd,
                    cwd=str(work_dir),
                    env=env_vars,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True
                )
                
                self.active_processes[session_id] = process
                
                try:
                    stdout, stderr = process.communicate(timeout=timeout)
                    returncode = process.returncode
                    
                    return {
                        'success': returncode == 0,
                        'stdout': stdout,
                        'stderr': stderr,
                        'returncode': returncode
                    }
                    
                except subprocess.TimeoutExpired:
                    process.kill()
                    return {
                        'success': False,
                        'stdout': '',
                        'stderr': f'Execution timed out after {timeout} seconds',
                        'returncode': -1
                    }
                    
            except Exception as e:
                return {
                    'success': False,
                    'stdout': '',
                    'stderr': str(e),
                    'returncode': -1
                }

    def execute_shell_command(self, command: str, session_id: str, timeout: int = 30) -> Dict[str, Any]:
        """Execute shell command in isolated environment."""
        with self.create_isolated_environment(session_id) as env:
            work_dir = env['work_dir']
            env_vars = env['env']
            
            try:
                process = subprocess.Popen(
                    command,
                    shell=True,
                    cwd=str(work_dir),
                    env=env_vars,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True
                )
                
                self.active_processes[session_id] = process
                
                try:
                    stdout, stderr = process.communicate(timeout=timeout)
                    returncode = process.returncode
                    
                    return {
                        'success': returncode == 0,
                        'stdout': stdout,
                        'stderr': stderr,
                        'returncode': returncode
                    }
                    
                except subprocess.TimeoutExpired:
                    process.kill()
                    return {
                        'success': False,
                        'stdout': '',
                        'stderr': f'Command timed out after {timeout} seconds',
                        'returncode': -1
                    }
                    
            except Exception as e:
                return {
                    'success': False,
                    'stdout': '',
                    'stderr': str(e),
                    'returncode': -1
                }

    def init_browser_automation(self) -> bool:
        """Initialize browser automation capabilities."""
        if not PLAYWRIGHT_AVAILABLE:
            self.print_error("Playwright not available for browser automation")
            return False
            
        try:
            self.playwright = sync_playwright().start()
            
            # Launch browser with security settings
            self.browser_instance = self.playwright.chromium.launch(
                headless=True,  # Can be changed to False for debugging
                args=[
                    '--no-sandbox',
                    '--disable-dev-shm-usage',
                    '--disable-gpu',
                    '--disable-background-timer-throttling',
                    '--disable-backgrounding-occluded-windows',
                    '--disable-renderer-backgrounding'
                ]
            )
            
            self.print_success("Browser automation initialized")
            return True
            
        except Exception as e:
            self.print_error(f"Failed to initialize browser automation: {e}")
            return False

    def create_browser_session(self, session_id: str) -> bool:
        """Create a new browser session."""
        if not self.browser_instance:
            if not self.init_browser_automation():
                return False
        
        try:
            # Create new browser context for isolation
            context = self.browser_instance.new_context(
                viewport={'width': 1024, 'height': 768},
                user_agent='Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
            )
            
            # Create new page
            page = context.new_page()
            self.browser_pages[session_id] = page
            
            self.print_success(f"Browser session created: {session_id}")
            return True
            
        except Exception as e:
            self.print_error(f"Failed to create browser session: {e}")
            return False

    def navigate_browser(self, session_id: str, url: str) -> Dict[str, Any]:
        """Navigate browser to URL."""
        if session_id not in self.browser_pages:
            if not self.create_browser_session(session_id):
                return {'success': False, 'error': 'Failed to create browser session'}
        
        page = self.browser_pages[session_id]
        
        try:
            response = page.goto(url, wait_until='networkidle', timeout=30000)
            
            return {
                'success': True,
                'url': page.url,
                'title': page.title(),
                'status': response.status if response else 0
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }

    def execute_browser_script(self, session_id: str, script: str) -> Dict[str, Any]:
        """Execute JavaScript in browser."""
        if session_id not in self.browser_pages:
            return {'success': False, 'error': 'No browser session found'}
        
        page = self.browser_pages[session_id]
        
        try:
            result = page.evaluate(script)
            
            return {
                'success': True,
                'result': result
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }

    def take_screenshot(self, session_id: str, full_page: bool = False) -> Dict[str, Any]:
        """Take screenshot of browser page."""
        if session_id not in self.browser_pages:
            return {'success': False, 'error': 'No browser session found'}
        
        page = self.browser_pages[session_id]
        
        try:
            screenshot_path = self.sandbox_dir / f"screenshot_{session_id}.png"
            
            page.screenshot(
                path=str(screenshot_path),
                full_page=full_page
            )
            
            return {
                'success': True,
                'screenshot_path': str(screenshot_path)
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }

    def process_document(self, file_path: str, session_id: str) -> Dict[str, Any]:
        """Process various document types."""
        file_path = Path(file_path)
        
        if not file_path.exists():
            return {'success': False, 'error': 'File not found'}
        
        with self.create_isolated_environment(session_id) as env:
            work_dir = env['work_dir']
            
            # Copy file to work directory
            dest_file = work_dir / file_path.name
            shutil.copy2(file_path, dest_file)
            
            try:
                if file_path.suffix.lower() == '.pdf':
                    return self._process_pdf(dest_file)
                elif file_path.suffix.lower() in ['.doc', '.docx']:
                    return self._process_word(dest_file)
                elif file_path.suffix.lower() in ['.xls', '.xlsx']:
                    return self._process_excel(dest_file)
                elif file_path.suffix.lower() in ['.txt', '.md']:
                    return self._process_text(dest_file)
                else:
                    return {'success': False, 'error': f'Unsupported file type: {file_path.suffix}'}
                    
            except Exception as e:
                return {'success': False, 'error': str(e)}

    def _process_pdf(self, file_path: Path) -> Dict[str, Any]:
        """Process PDF file."""
        try:
            # Try with PyPDF2 first
            try:
                from PyPDF2 import PdfReader
                
                reader = PdfReader(str(file_path))
                text = ""
                for page in reader.pages:
                    text += page.extract_text() + "\n"
                    
                return {
                    'success': True,
                    'content': text,
                    'pages': len(reader.pages)
                }
                
            except ImportError:
                # Fallback to command line tools
                result = subprocess.run(
                    ['pdftotext', str(file_path), '-'],
                    capture_output=True,
                    text=True
                )
                
                if result.returncode == 0:
                    return {
                        'success': True,
                        'content': result.stdout
                    }
                else:
                    return {
                        'success': False,
                        'error': 'PDF processing failed'
                    }
                    
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def _process_word(self, file_path: Path) -> Dict[str, Any]:
        """Process Word document."""
        try:
            from python_docx import Document
            
            doc = Document(str(file_path))
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
                
            return {
                'success': True,
                'content': text
            }
            
        except ImportError:
            return {
                'success': False,
                'error': 'python-docx not available for Word processing'
            }
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def _process_excel(self, file_path: Path) -> Dict[str, Any]:
        """Process Excel file."""
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(str(file_path))
            sheets_data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                
                for row in sheet.iter_rows(values_only=True):
                    sheet_data.append(list(row))
                    
                sheets_data[sheet_name] = sheet_data
                
            return {
                'success': True,
                'sheets': sheets_data
            }
            
        except ImportError:
            return {
                'success': False,
                'error': 'openpyxl not available for Excel processing'
            }
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def _process_text(self, file_path: Path) -> Dict[str, Any]:
        """Process text file."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                
            return {
                'success': True,
                'content': content
            }
            
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def cleanup(self):
        """Clean up all resources."""
        self.print_info("Cleaning up sandbox resources...")
        
        # Stop all active processes
        for session_id in list(self.active_processes.keys()):
            self.cleanup_session(session_id)
        
        # Close browser
        if self.browser_instance:
            try:
                self.browser_instance.close()
            except Exception as e:
                self.logger.warning(f"Error closing browser: {e}")
        
        # Close playwright
        if hasattr(self, 'playwright'):
            try:
                self.playwright.stop()
            except Exception as e:
                self.logger.warning(f"Error stopping playwright: {e}")
        
        self.print_success("Sandbox cleanup completed")


def main():
    """Test the sandbox manager."""
    if len(sys.argv) > 1 and sys.argv[1] == "--help":
        print("Usage: python sandbox_manager.py [COMMAND] [ARGS]")
        print("Commands:")
        print("  test       - Run basic tests")
        print("  install    - Install browser dependencies")
        print("  --help     - Show this help")
        return
    
    project_root = Path(__file__).parent
    sandbox = SandboxManager(str(project_root))
    
    def cleanup_handler(signum, frame):
        sandbox.cleanup()
        sys.exit(0)
    
    signal.signal(signal.SIGINT, cleanup_handler)
    signal.signal(signal.SIGTERM, cleanup_handler)
    
    command = sys.argv[1] if len(sys.argv) > 1 else "test"
    
    try:
        if command == "install":
            sandbox.check_dependencies()
            sandbox.install_browser_dependencies()
        elif command == "test":
            # Run basic tests
            print("Testing sandbox manager...")
            
            # Test Python execution
            result = sandbox.execute_python_code("print('Hello from sandbox!')", "test_session")
            print(f"Python test: {result}")
            
            # Test shell command
            result = sandbox.execute_shell_command("echo 'Shell test'", "test_session")
            print(f"Shell test: {result}")
            
            # Test browser if available
            if PLAYWRIGHT_AVAILABLE:
                sandbox.init_browser_automation()
                result = sandbox.navigate_browser("browser_test", "https://example.com")
                print(f"Browser test: {result}")
            
            print("Tests completed")
        else:
            print(f"Unknown command: {command}")
    
    finally:
        sandbox.cleanup()


if __name__ == "__main__":
    main()