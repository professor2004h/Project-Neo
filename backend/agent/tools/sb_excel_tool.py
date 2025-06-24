from typing import Optional, Dict, Any, List, Union
import json
import io
import base64
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.pivot.table import PivotTable, PivotField
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from agentpress.tool import ToolResult, openapi_schema, xml_schema
from sandbox.tool_base import SandboxToolsBase
from agentpress.thread_manager import ThreadManager
from utils.logger import logger

class SandboxExcelTool(SandboxToolsBase):
    """Tool for comprehensive Excel file manipulation in a Daytona sandbox using openpyxl."""

    def __init__(self, project_id: str, thread_manager: ThreadManager):
        super().__init__(project_id, thread_manager)
        self.workspace_path = "/workspace"
        self._workbooks: Dict[str, Workbook] = {}  # Cache for open workbooks

    def clean_path(self, path: str) -> str:
        """Clean and normalize a path to be relative to /workspace"""
        return super().clean_path(path)

    def _file_exists(self, path: str) -> bool:
        """Check if a file exists in the sandbox"""
        try:
            self.sandbox.fs.get_file_info(path)
            return True
        except Exception:
            return False

    async def _ensure_openpyxl_installed(self):
        """Ensure openpyxl is installed in the sandbox"""
        try:
            # Check if openpyxl is available
            response = self.sandbox.process.exec("python3 -c 'import openpyxl; print(openpyxl.__version__)'", timeout=10)
            if response.exit_code != 0:
                # Install openpyxl if not available (fallback for older containers)
                logger.info("Installing openpyxl and pandas in sandbox...")
                install_response = self.sandbox.process.exec("pip install --no-cache-dir openpyxl==3.1.5 pandas==2.2.3", timeout=120)
                if install_response.exit_code != 0:
                    raise Exception(f"Failed to install openpyxl: {install_response.result}")
                logger.info("Successfully installed openpyxl and pandas")
            else:
                logger.debug("openpyxl already available in sandbox")
                
            # Also verify pandas is available
            pandas_response = self.sandbox.process.exec("python3 -c 'import pandas; print(pandas.__version__)'", timeout=10)
            if pandas_response.exit_code != 0:
                logger.info("Installing pandas in sandbox...")
                install_response = self.sandbox.process.exec("pip install --no-cache-dir pandas==2.2.3", timeout=120)
                if install_response.exit_code != 0:
                    raise Exception(f"Failed to install pandas: {install_response.result}")
                    
        except Exception as e:
            logger.warning(f"Could not verify openpyxl/pandas installation: {e}")
            # Don't raise the exception - let the tool try to work anyway
            # The actual tool methods will fail with clearer error messages if imports don't work

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "create_workbook",
            "description": "Create a new Excel workbook (.xlsx) at the specified path. The path must be relative to /workspace.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path where the Excel file will be created, relative to /workspace (e.g., 'data/report.xlsx')"
                    },
                    "sheet_names": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Optional list of sheet names to create. If not provided, creates default 'Sheet1'",
                        "default": ["Sheet1"]
                    }
                },
                "required": ["file_path"]
            }
        }
    })
    @xml_schema(
        tag_name="create-workbook",
        mappings=[
            {"param_name": "file_path", "node_type": "attribute", "path": "."},
            {"param_name": "sheet_names", "node_type": "element", "path": "sheet_names", "required": False}
        ],
        example='''
        <function_calls>
        <invoke name="create_workbook">
        <parameter name="file_path">data/sales_report.xlsx</parameter>
        <parameter name="sheet_names">["Sales", "Summary", "Charts"]</parameter>
        </invoke>
        </function_calls>
        '''
    )
    async def create_workbook(self, file_path: str, sheet_names: Optional[List[str]] = None) -> ToolResult:
        try:
            await self._ensure_sandbox()
            await self._ensure_openpyxl_installed()
            
            file_path = self.clean_path(file_path)
            full_path = f"{self.workspace_path}/{file_path}"
            
            if self._file_exists(full_path):
                return self.fail_response(f"File '{file_path}' already exists. Use open_workbook to modify existing files.")
            
            # Create new workbook
            wb = Workbook()
            
            # Remove default sheet if custom sheets are specified
            if sheet_names:
                # Remove default sheet
                default_sheet = wb.active
                wb.remove(default_sheet)
                
                # Add custom sheets
                for sheet_name in sheet_names:
                    wb.create_sheet(title=sheet_name)
            
            # Save workbook to bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Create parent directories if needed
            parent_dir = '/'.join(full_path.split('/')[:-1])
            if parent_dir:
                self.sandbox.fs.create_folder(parent_dir, "755")
            
            # Upload to sandbox
            self.sandbox.fs.upload_file(buffer.getvalue(), full_path)
            
            # Cache the workbook
            self._workbooks[file_path] = wb
            
            sheets_info = [sheet.title for sheet in wb.worksheets]
            return self.success_response(f"Excel workbook '{file_path}' created successfully with sheets: {sheets_info}")
            
        except Exception as e:
            return self.fail_response(f"Error creating workbook: {str(e)}")

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "open_workbook",
            "description": "Open an existing Excel workbook for editing. The file path must be relative to /workspace.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file to open, relative to /workspace (e.g., 'data/report.xlsx')"
                    }
                },
                "required": ["file_path"]
            }
        }
    })
    @xml_schema(
        tag_name="open-workbook",
        mappings=[
            {"param_name": "file_path", "node_type": "attribute", "path": "."}
        ],
        example='''
        <function_calls>
        <invoke name="open_workbook">
        <parameter name="file_path">data/sales_report.xlsx</parameter>
        </invoke>
        </function_calls>
        '''
    )
    async def open_workbook(self, file_path: str) -> ToolResult:
        try:
            await self._ensure_sandbox()
            await self._ensure_openpyxl_installed()
            
            file_path = self.clean_path(file_path)
            full_path = f"{self.workspace_path}/{file_path}"
            
            if not self._file_exists(full_path):
                return self.fail_response(f"File '{file_path}' does not exist")
            
            # Download file content
            file_content = self.sandbox.fs.download_file(full_path)
            
            # Load workbook from bytes
            buffer = io.BytesIO(file_content)
            wb = load_workbook(buffer)
            
            # Cache the workbook
            self._workbooks[file_path] = wb
            
            sheets_info = [sheet.title for sheet in wb.worksheets]
            return self.success_response(f"Excel workbook '{file_path}' opened successfully. Sheets: {sheets_info}")
            
        except Exception as e:
            return self.fail_response(f"Error opening workbook: {str(e)}")

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "save_workbook",
            "description": "Save an open workbook to the file system. The workbook must be previously opened or created.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file to save, relative to /workspace"
                    }
                },
                "required": ["file_path"]
            }
        }
    })
    @xml_schema(
        tag_name="save-workbook",
        mappings=[
            {"param_name": "file_path", "node_type": "attribute", "path": "."}
        ],
        example='''
        <function_calls>
        <invoke name="save_workbook">
        <parameter name="file_path">data/sales_report.xlsx</parameter>
        </invoke>
        </function_calls>
        '''
    )
    async def save_workbook(self, file_path: str) -> ToolResult:
        try:
            await self._ensure_sandbox()
            
            file_path = self.clean_path(file_path)
            
            if file_path not in self._workbooks:
                return self.fail_response(f"Workbook '{file_path}' is not open. Use open_workbook or create_workbook first.")
            
            wb = self._workbooks[file_path]
            full_path = f"{self.workspace_path}/{file_path}"
            
            # Save workbook to bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Upload to sandbox
            self.sandbox.fs.upload_file(buffer.getvalue(), full_path)
            
            return self.success_response(f"Workbook '{file_path}' saved successfully")
            
        except Exception as e:
            return self.fail_response(f"Error saving workbook: {str(e)}")

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "write_cell_data",
            "description": "Write data to specific cells in a worksheet. Supports writing single values, ranges, or bulk data.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file, relative to /workspace"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the worksheet to write to"
                    },
                    "cell_range": {
                        "type": "string",
                        "description": "Cell range to write to (e.g., 'A1', 'A1:C3', 'B2:D10')"
                    },
                    "data": {
                        "type": "array",
                        "items": {"type": "array"},
                        "description": "2D array of data to write. For single cell, use [[value]]. For row, use [[val1, val2, val3]]. For multiple rows, use [[row1_val1, row1_val2], [row2_val1, row2_val2]]"
                    }
                },
                "required": ["file_path", "sheet_name", "cell_range", "data"]
            }
        }
    })
    @xml_schema(
        tag_name="write-cell-data",
        mappings=[
            {"param_name": "file_path", "node_type": "attribute", "path": "."},
            {"param_name": "sheet_name", "node_type": "attribute", "path": "sheet_name"},
            {"param_name": "cell_range", "node_type": "attribute", "path": "cell_range"},
            {"param_name": "data", "node_type": "element", "path": "data"}
        ],
        example='''
        <function_calls>
        <invoke name="write_cell_data">
        <parameter name="file_path">data/sales_report.xlsx</parameter>
        <parameter name="sheet_name">Sales</parameter>
        <parameter name="cell_range">A1:C2</parameter>
        <parameter name="data">[["Name", "Sales", "Commission"], ["John", 1000, 100]]</parameter>
        </invoke>
        </function_calls>
        '''
    )
    async def write_cell_data(self, file_path: str, sheet_name: str, cell_range: str, data: List[List[Any]]) -> ToolResult:
        try:
            await self._ensure_sandbox()
            
            file_path = self.clean_path(file_path)
            
            if file_path not in self._workbooks:
                # Try to open the workbook
                result = await self.open_workbook(file_path)
                if not result.success:
                    return result
            
            wb = self._workbooks[file_path]
            
            # Get or create worksheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
            
            # Parse cell range
            if ':' in cell_range:
                # Range of cells
                start_cell, end_cell = cell_range.split(':')
                start_row = int(''.join(filter(str.isdigit, start_cell)))
                start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))
                
                # Write data row by row
                for row_idx, row_data in enumerate(data):
                    for col_idx, value in enumerate(row_data):
                        cell = ws.cell(row=start_row + row_idx, column=start_col + col_idx)
                        cell.value = value
            else:
                # Single cell
                cell = ws[cell_range]
                if data and data[0]:
                    cell.value = data[0][0]
            
            return self.success_response(f"Data written to {cell_range} in sheet '{sheet_name}'")
            
        except Exception as e:
            return self.fail_response(f"Error writing cell data: {str(e)}")

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "read_cell_data", 
            "description": "Read data from specific cells or ranges in a worksheet.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file, relative to /workspace"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the worksheet to read from"
                    },
                    "cell_range": {
                        "type": "string",
                        "description": "Cell range to read (e.g., 'A1', 'A1:C10', 'B:B' for entire column)"
                    }
                },
                "required": ["file_path", "sheet_name", "cell_range"]
            }
        }
    })
    @xml_schema(
        tag_name="read-cell-data",
        mappings=[
            {"param_name": "file_path", "node_type": "attribute", "path": "."},
            {"param_name": "sheet_name", "node_type": "attribute", "path": "sheet_name"},
            {"param_name": "cell_range", "node_type": "attribute", "path": "cell_range"}
        ],
        example='''
        <function_calls>
        <invoke name="read_cell_data">
        <parameter name="file_path">data/sales_report.xlsx</parameter>
        <parameter name="sheet_name">Sales</parameter>
        <parameter name="cell_range">A1:C10</parameter>
        </invoke>
        </function_calls>
        '''
    )
    async def read_cell_data(self, file_path: str, sheet_name: str, cell_range: str) -> ToolResult:
        try:
            await self._ensure_sandbox()
            
            file_path = self.clean_path(file_path)
            
            if file_path not in self._workbooks:
                result = await self.open_workbook(file_path)
                if not result.success:
                    return result
            
            wb = self._workbooks[file_path]
            
            if sheet_name not in wb.sheetnames:
                return self.fail_response(f"Sheet '{sheet_name}' does not exist")
            
            ws = wb[sheet_name]
            
            # Read data from range
            data = []
            for row in ws[cell_range]:
                if isinstance(row, tuple):
                    row_data = [cell.value for cell in row]
                else:
                    row_data = [row.value]
                data.append(row_data)
            
            return self.success_response({
                "data": data,
                "range": cell_range,
                "sheet": sheet_name,
                "rows": len(data),
                "columns": len(data[0]) if data else 0
            })
            
        except Exception as e:
            return self.fail_response(f"Error reading cell data: {str(e)}")

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "format_cells",
            "description": "Apply formatting to cells including font, alignment, borders, and fill colors.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file, relative to /workspace"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the worksheet"
                    },
                    "cell_range": {
                        "type": "string",
                        "description": "Cell range to format (e.g., 'A1:C10')"
                    },
                    "font_name": {
                        "type": "string",
                        "description": "Font name (e.g., 'Arial', 'Times New Roman')",
                        "default": None
                    },
                    "font_size": {
                        "type": "integer",
                        "description": "Font size",
                        "default": None
                    },
                    "bold": {
                        "type": "boolean",
                        "description": "Make text bold",
                        "default": None
                    },
                    "italic": {
                        "type": "boolean",
                        "description": "Make text italic",
                        "default": None
                    },
                    "font_color": {
                        "type": "string",
                        "description": "Font color in hex format (e.g., 'FF0000' for red)",
                        "default": None
                    },
                    "background_color": {
                        "type": "string",
                        "description": "Background color in hex format (e.g., 'FFFF00' for yellow)",
                        "default": None
                    },
                    "horizontal_alignment": {
                        "type": "string",
                        "enum": ["left", "center", "right"],
                        "description": "Horizontal alignment",
                        "default": None
                    },
                    "vertical_alignment": {
                        "type": "string",
                        "enum": ["top", "center", "bottom"],
                        "description": "Vertical alignment",
                        "default": None
                    },
                    "border_style": {
                        "type": "string",
                        "enum": ["thin", "thick", "medium"],
                        "description": "Border style",
                        "default": None
                    }
                },
                "required": ["file_path", "sheet_name", "cell_range"]
            }
        }
    })
    @xml_schema(
        tag_name="format-cells",
        mappings=[
            {"param_name": "file_path", "node_type": "attribute", "path": "."},
            {"param_name": "sheet_name", "node_type": "attribute", "path": "sheet_name"},
            {"param_name": "cell_range", "node_type": "attribute", "path": "cell_range"},
            {"param_name": "font_name", "node_type": "attribute", "path": ".", "required": False},
            {"param_name": "font_size", "node_type": "attribute", "path": ".", "required": False},
            {"param_name": "bold", "node_type": "attribute", "path": ".", "required": False},
            {"param_name": "italic", "node_type": "attribute", "path": ".", "required": False},
            {"param_name": "font_color", "node_type": "attribute", "path": ".", "required": False},
            {"param_name": "background_color", "node_type": "attribute", "path": ".", "required": False},
            {"param_name": "horizontal_alignment", "node_type": "attribute", "path": ".", "required": False},
            {"param_name": "vertical_alignment", "node_type": "attribute", "path": ".", "required": False},
            {"param_name": "border_style", "node_type": "attribute", "path": ".", "required": False}
        ],
        example='''
        <function_calls>
        <invoke name="format_cells">
        <parameter name="file_path">data/sales_report.xlsx</parameter>
        <parameter name="sheet_name">Sales</parameter>
        <parameter name="cell_range">A1:C1</parameter>
        <parameter name="bold">true</parameter>
        <parameter name="background_color">CCCCCC</parameter>
        <parameter name="horizontal_alignment">center</parameter>
        <parameter name="border_style">thin</parameter>
        </invoke>
        </function_calls>
        '''
    )
    async def format_cells(self, file_path: str, sheet_name: str, cell_range: str, 
                          font_name: Optional[str] = None, font_size: Optional[int] = None,
                          bold: Optional[bool] = None, italic: Optional[bool] = None,
                          font_color: Optional[str] = None, background_color: Optional[str] = None,
                          horizontal_alignment: Optional[str] = None, vertical_alignment: Optional[str] = None,
                          border_style: Optional[str] = None) -> ToolResult:
        try:
            await self._ensure_sandbox()
            
            file_path = self.clean_path(file_path)
            
            if file_path not in self._workbooks:
                result = await self.open_workbook(file_path)
                if not result.success:
                    return result
            
            wb = self._workbooks[file_path]
            
            if sheet_name not in wb.sheetnames:
                return self.fail_response(f"Sheet '{sheet_name}' does not exist")
            
            ws = wb[sheet_name]
            
            # Apply formatting to range
            for row in ws[cell_range]:
                cells = row if isinstance(row, tuple) else [row]
                for cell in cells:
                    # Font formatting
                    if any([font_name, font_size, bold, italic, font_color]):
                        font_kwargs = {}
                        if font_name:
                            font_kwargs['name'] = font_name
                        if font_size:
                            font_kwargs['size'] = font_size
                        if bold is not None:
                            font_kwargs['bold'] = bold
                        if italic is not None:
                            font_kwargs['italic'] = italic
                        if font_color:
                            font_kwargs['color'] = font_color
                        
                        cell.font = Font(**font_kwargs)
                    
                    # Background fill
                    if background_color:
                        cell.fill = PatternFill(start_color=background_color, end_color=background_color, fill_type='solid')
                    
                    # Alignment
                    if horizontal_alignment or vertical_alignment:
                        align_kwargs = {}
                        if horizontal_alignment:
                            align_kwargs['horizontal'] = horizontal_alignment
                        if vertical_alignment:
                            align_kwargs['vertical'] = vertical_alignment
                        cell.alignment = Alignment(**align_kwargs)
                    
                    # Borders
                    if border_style:
                        side = Side(style=border_style)
                        cell.border = Border(left=side, right=side, top=side, bottom=side)
            
            return self.success_response(f"Formatting applied to {cell_range} in sheet '{sheet_name}'")
            
        except Exception as e:
            return self.fail_response(f"Error formatting cells: {str(e)}")

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "create_chart",
            "description": "Create a chart in the worksheet from data range. Supports bar, line, and pie charts.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file, relative to /workspace"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the worksheet containing the data"
                    },
                    "chart_type": {
                        "type": "string",
                        "enum": ["bar", "line", "pie"],
                        "description": "Type of chart to create"
                    },
                    "data_range": {
                        "type": "string",
                        "description": "Range containing the data for the chart (e.g., 'A1:B10')"
                    },
                    "chart_title": {
                        "type": "string",
                        "description": "Title for the chart",
                        "default": "Chart"
                    },
                    "position": {
                        "type": "string",
                        "description": "Cell position where chart will be placed (e.g., 'E1')",
                        "default": "E1"
                    }
                },
                "required": ["file_path", "sheet_name", "chart_type", "data_range"]
            }
        }
    })
    @xml_schema(
        tag_name="create-chart",
        mappings=[
            {"param_name": "file_path", "node_type": "attribute", "path": "."},
            {"param_name": "sheet_name", "node_type": "attribute", "path": "sheet_name"},
            {"param_name": "chart_type", "node_type": "attribute", "path": "chart_type"},
            {"param_name": "data_range", "node_type": "attribute", "path": "data_range"},
            {"param_name": "chart_title", "node_type": "attribute", "path": ".", "required": False},
            {"param_name": "position", "node_type": "attribute", "path": ".", "required": False}
        ],
        example='''
        <function_calls>
        <invoke name="create_chart">
        <parameter name="file_path">data/sales_report.xlsx</parameter>
        <parameter name="sheet_name">Sales</parameter>
        <parameter name="chart_type">bar</parameter>
        <parameter name="data_range">A1:B10</parameter>
        <parameter name="chart_title">Monthly Sales</parameter>
        <parameter name="position">D1</parameter>
        </invoke>
        </function_calls>
        '''
    )
    async def create_chart(self, file_path: str, sheet_name: str, chart_type: str, data_range: str,
                          chart_title: str = "Chart", position: str = "E1") -> ToolResult:
        try:
            await self._ensure_sandbox()
            
            file_path = self.clean_path(file_path)
            
            if file_path not in self._workbooks:
                result = await self.open_workbook(file_path)
                if not result.success:
                    return result
            
            wb = self._workbooks[file_path]
            
            if sheet_name not in wb.sheetnames:
                return self.fail_response(f"Sheet '{sheet_name}' does not exist")
            
            ws = wb[sheet_name]
            
            # Create reference to data
            data_ref = Reference(ws, range_string=f"{sheet_name}!{data_range}")
            
            # Create chart based on type
            if chart_type == "bar":
                chart = BarChart()
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            else:
                return self.fail_response(f"Unsupported chart type: {chart_type}")
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.title = chart_title
            
            # Add chart to worksheet
            ws.add_chart(chart, position)
            
            return self.success_response(f"{chart_type.capitalize()} chart '{chart_title}' created at {position} in sheet '{sheet_name}'")
            
        except Exception as e:
            return self.fail_response(f"Error creating chart: {str(e)}")

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "list_worksheets",
            "description": "List all worksheets in an Excel workbook.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file, relative to /workspace"
                    }
                },
                "required": ["file_path"]
            }
        }
    })
    @xml_schema(
        tag_name="list-worksheets",
        mappings=[
            {"param_name": "file_path", "node_type": "attribute", "path": "."}
        ],
        example='''
        <function_calls>
        <invoke name="list_worksheets">
        <parameter name="file_path">data/sales_report.xlsx</parameter>
        </invoke>
        </function_calls>
        '''
    )
    async def list_worksheets(self, file_path: str) -> ToolResult:
        try:
            await self._ensure_sandbox()
            
            file_path = self.clean_path(file_path)
            
            if file_path not in self._workbooks:
                result = await self.open_workbook(file_path)
                if not result.success:
                    return result
            
            wb = self._workbooks[file_path]
            worksheets = [sheet.title for sheet in wb.worksheets]
            
            return self.success_response({
                "worksheets": worksheets,
                "count": len(worksheets)
            })
            
        except Exception as e:
            return self.fail_response(f"Error listing worksheets: {str(e)}")

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "create_worksheet",
            "description": "Create a new worksheet in an existing workbook.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file, relative to /workspace"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name for the new worksheet"
                    },
                    "index": {
                        "type": "integer",
                        "description": "Position to insert the sheet (0-based). If not specified, adds at the end",
                        "default": None
                    }
                },
                "required": ["file_path", "sheet_name"]
            }
        }
    })
    @xml_schema(
        tag_name="create-worksheet",
        mappings=[
            {"param_name": "file_path", "node_type": "attribute", "path": "."},
            {"param_name": "sheet_name", "node_type": "attribute", "path": "sheet_name"},
            {"param_name": "index", "node_type": "attribute", "path": ".", "required": False}
        ],
        example='''
        <function_calls>
        <invoke name="create_worksheet">
        <parameter name="file_path">data/sales_report.xlsx</parameter>
        <parameter name="sheet_name">Q4 Summary</parameter>
        <parameter name="index">1</parameter>
        </invoke>
        </function_calls>
        '''
    )
    async def create_worksheet(self, file_path: str, sheet_name: str, index: Optional[int] = None) -> ToolResult:
        try:
            await self._ensure_sandbox()
            
            file_path = self.clean_path(file_path)
            
            if file_path not in self._workbooks:
                result = await self.open_workbook(file_path)
                if not result.success:
                    return result
            
            wb = self._workbooks[file_path]
            
            if sheet_name in wb.sheetnames:
                return self.fail_response(f"Worksheet '{sheet_name}' already exists")
            
            # Create new worksheet
            if index is not None:
                ws = wb.create_sheet(title=sheet_name, index=index)
            else:
                ws = wb.create_sheet(title=sheet_name)
            
            return self.success_response(f"Worksheet '{sheet_name}' created successfully")
            
        except Exception as e:
            return self.fail_response(f"Error creating worksheet: {str(e)}")

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "delete_worksheet",
            "description": "Delete a worksheet from the workbook.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file, relative to /workspace"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the worksheet to delete"
                    }
                },
                "required": ["file_path", "sheet_name"]
            }
        }
    })
    @xml_schema(
        tag_name="delete-worksheet",
        mappings=[
            {"param_name": "file_path", "node_type": "attribute", "path": "."},
            {"param_name": "sheet_name", "node_type": "attribute", "path": "sheet_name"}
        ],
        example='''
        <function_calls>
        <invoke name="delete_worksheet">
        <parameter name="file_path">data/sales_report.xlsx</parameter>
        <parameter name="sheet_name">Temp Sheet</parameter>
        </invoke>
        </function_calls>
        '''
    )
    async def delete_worksheet(self, file_path: str, sheet_name: str) -> ToolResult:
        try:
            await self._ensure_sandbox()
            
            file_path = self.clean_path(file_path)
            
            if file_path not in self._workbooks:
                result = await self.open_workbook(file_path)
                if not result.success:
                    return result
            
            wb = self._workbooks[file_path]
            
            if sheet_name not in wb.sheetnames:
                return self.fail_response(f"Worksheet '{sheet_name}' does not exist")
            
            if len(wb.worksheets) <= 1:
                return self.fail_response("Cannot delete the last remaining worksheet")
            
            # Delete worksheet
            ws = wb[sheet_name]
            wb.remove(ws)
            
            return self.success_response(f"Worksheet '{sheet_name}' deleted successfully")
            
        except Exception as e:
            return self.fail_response(f"Error deleting worksheet: {str(e)}")

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "close_workbook",
            "description": "Close a workbook and remove it from memory. This does not save the workbook - use save_workbook first if needed.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file to close, relative to /workspace"
                    }
                },
                "required": ["file_path"]
            }
        }
    })
    @xml_schema(
        tag_name="close-workbook",
        mappings=[
            {"param_name": "file_path", "node_type": "attribute", "path": "."}
        ],
        example='''
        <function_calls>
        <invoke name="close_workbook">
        <parameter name="file_path">data/sales_report.xlsx</parameter>
        </invoke>
        </function_calls>
        '''
    )
    async def close_workbook(self, file_path: str) -> ToolResult:
        try:
            file_path = self.clean_path(file_path)
            
            if file_path in self._workbooks:
                del self._workbooks[file_path]
                return self.success_response(f"Workbook '{file_path}' closed successfully")
            else:
                return self.success_response(f"Workbook '{file_path}' was not open")
            
        except Exception as e:
            return self.fail_response(f"Error closing workbook: {str(e)}") 